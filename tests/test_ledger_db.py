# -*- coding: utf-8 -*-
"""台账数据层（database/ledger_db.py）回归测试

覆盖 insert/update/delete、分页筛选（分类/类别/署名/关键词）、
类别候选合并、按署名统计、按分类分 sheet 导出。

隔离方式同 test_aftersale_batch_ops：tmp SQLite + monkeypatch，
不触碰真实 tables.db；MySQL 开关强制关闭保证走 SQLite 方言。
"""
import sqlite3

import pytest

import database.backend as backend
import database.table_db as table_db
import database.ledger_db as ldb
from database import schema


@pytest.fixture
def db(monkeypatch, tmp_path):
    """临时库：建 ledger_records 表，关闭 MySQL 测试开关"""
    path = str(tmp_path / "t.db")
    monkeypatch.setattr(table_db, "DB_PATH", path)
    monkeypatch.setattr(table_db, "_conn", None)
    monkeypatch.setattr(table_db, "_ensure_initialized", lambda c: None)
    monkeypatch.setattr(backend, "is_mysql_test_mode", lambda: False)
    sl = sqlite3.connect(path)
    sl.executescript(schema.to_sqlite_ddl("ledger_records"))
    sl.commit()
    sl.close()
    return path


def _rec(**kw):
    """构造一条完整记录（必填三项 + 选填默认空）"""
    base = {
        "category": "问题", "kind": "遮挡问题", "room_name": "9601",
        "video_name": "", "frame": "", "description": "",
        "repro": "", "new_program": "", "remark": "", "signer": "沈喆",
    }
    base.update(kw)
    return base


# ==================== 增删改 ====================

def test_insert_record_fills_timestamps(db):
    rid = ldb.insert_record(_rec(
        created_at="1999-01-01 00:00:00",  # 调用方传入时间应被忽略
        video_name="v1.mp4", frame="400"))
    assert rid > 0
    sl = sqlite3.connect(db)
    row = sl.execute(
        "SELECT created_at, updated_at FROM ledger_records WHERE id=?",
        (rid,)).fetchone()
    sl.close()
    assert row[0] != "1999-01-01 00:00:00"  # 服务端统一维护
    assert row[0] == row[1]
    assert len(str(row[0])) >= 16            # YYYY-MM-DD HH:MM:SS


def test_update_record_only_touches_fields(db):
    rid = ldb.insert_record(_rec(video_name="a.mp4"))
    ok = ldb.update_record(rid, {"kind": "丢球:袋口", "remark": "改备注"})
    assert ok
    sl = sqlite3.connect(db)
    row = sl.execute(
        "SELECT kind, remark, video_name, updated_at FROM ledger_records "
        "WHERE id=?", (rid,)).fetchone()
    sl.close()
    assert row[0] == "丢球:袋口"
    assert row[1] == "改备注"
    assert row[2] == "a.mp4"      # 未传字段不误改
    assert row[3]                 # updated_at 已刷新


def test_update_record_missing_id(db):
    assert ldb.update_record(99999, {"kind": "x"}) is False


def test_delete_record(db):
    rid = ldb.insert_record(_rec())
    assert ldb.delete_record(rid) is True
    assert ldb.delete_record(rid) is False  # 已删不存在


# ==================== 分页筛选 ====================

def _seed(db):
    """3 条记录：不同分类/类别/署名，验证筛选与排序"""
    ldb.insert_record(_rec(category="问题", kind="遮挡问题", signer="甲",
                           video_name="v1"))
    ldb.insert_record(_rec(category="精度", kind="轻贴:不动", signer="乙",
                           video_name="v2", frame="500"))
    ldb.insert_record(_rec(category="使用", kind="让杆/换人", signer="甲",
                           description="开局前让杆"))


def test_query_page_all_and_order_desc(db):
    _seed(db)
    total, rows = ldb.query_page(1, 50)
    assert total == 3 and len(rows) == 3
    ids = [r["id"] for r in rows]
    assert ids == sorted(ids, reverse=True)  # 最新在前


def test_query_page_category_filter(db):
    _seed(db)
    total, rows = ldb.query_page(1, 50, category="精度")
    assert total == 1
    assert rows[0]["kind"] == "轻贴:不动"


def test_query_page_kind_and_signer_filter(db):
    _seed(db)
    total, rows = ldb.query_page(1, 50, kind="让杆/换人", signer="甲")
    assert total == 1
    assert rows[0]["category"] == "使用"


def test_query_page_keyword_search(db):
    _seed(db)
    # 署名命中两条（甲：问题 + 使用）
    total, rows = ldb.query_page(1, 50, keyword="甲")
    assert total == 2
    assert {r["signer"] for r in rows} == {"甲"}
    # 描述命中一条（OR 条件同行多字段命中不叠加）
    total2, rows2 = ldb.query_page(1, 50, keyword="开局前")
    assert total2 == 1
    assert rows2[0]["category"] == "使用"
    # kind 命中：让杆/换人 与 description 属同一行，仍只算一条
    total3, rows3 = ldb.query_page(1, 50, keyword="让杆")
    assert total3 == 1
    assert rows3[0]["kind"] == "让杆/换人"


def test_query_page_pagination(db):
    _seed(db)
    total, page1 = ldb.query_page(1, 2)
    assert total == 3 and len(page1) == 2
    _t, page2 = ldb.query_page(2, 2)
    assert len(page2) == 1
    assert {r["id"] for r in page1} & {r["id"] for r in page2} == set()


# ==================== 类别候选 ====================

def test_get_kind_candidates_preset_plus_free_input(db):
    rid = ldb.insert_record(_rec(category="精度", kind="自定义精度项"))
    cands = ldb.get_kind_candidates("精度")
    assert cands[:7] == list(ldb.KIND_CANDIDATES["精度"])  # 模板预置在前
    assert "自定义精度项" in cands                     # 库中自由输入合并
    assert cands.index("自定义精度项") > 6             # 追加在预置之后
    assert cands.count("自定义精度项") == 1            # 去重
    assert rid > 0


def test_get_kind_candidates_unknown_category(db):
    assert ldb.get_kind_candidates("不存在分类") == []


# ==================== 按署名统计 ====================

def test_stats_by_signer(db):
    _seed(db)
    ldb.insert_record(_rec(category="问题", kind="丢杆", signer="乙"))
    ldb.insert_record(_rec(category="问题", kind="丢球:袋口", signer=""))
    stats = ldb.stats_by_signer()
    by = {s["signer"]: s for s in stats}
    assert by["甲"]["问题"] == 1 and by["甲"]["使用"] == 1
    assert by["甲"]["total"] == 2
    assert by["乙"]["问题"] == 1 and by["乙"]["精度"] == 1
    assert by["乙"]["total"] == 2
    assert by["未署名"]["问题"] == 1 and by["未署名"]["total"] == 1
    # 每个 bucket 四分类键齐全
    for s in stats:
        assert set(ldb.CATEGORIES) <= set(s.keys())
    # 按 total 降序（甲/乙并列 2，未署名 1 在最后）
    assert stats[-1]["signer"] == "未署名"


def test_stats_by_signer_empty(db):
    assert ldb.stats_by_signer() == []


# ==================== 导出 xlsx ====================

def test_export_xlsx_sheets_and_rows(db, tmp_path):
    _seed(db)
    out = str(tmp_path / "ledger_export.xlsx")
    n = ldb.export_xlsx(out)
    assert n == 3
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert set(wb.sheetnames) == set(ldb.CATEGORIES)  # 四分类分 sheet
    ws = wb["问题"]
    headers = [c.value for c in ws[1]]
    assert headers == ["类别", "球房", "视频名", "帧数", "描述",
                       "复现", "新程序", "备注", "署名"]  # 与在线模板同结构
    assert ws.max_row == 2  # 表头 + 1 条问题记录


def test_export_xlsx_single_category(db, tmp_path):
    _seed(db)
    out = str(tmp_path / "single.xlsx")
    n = ldb.export_xlsx(out, category="使用")
    assert n == 1
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == ["使用"]
    assert wb["使用"].max_row == 2
