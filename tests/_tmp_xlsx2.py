# -*- coding: utf-8 -*-
"""临时：提取模板各类别取值分布 + 相机 sheet 完整内容（验证后删除）"""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r"C:\Users\shen_zhe\Desktop\在线模板.xlsx", data_only=True)

# 相机 sheet 全部内容
ws = wb["相机"]
print("== 相机 sheet 全内容 ==")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    print(row)

# 计数 sheet
ws = wb["计数"]
print("\n== 计数 sheet ==")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    print(row)

# 各 sheet 类别列取值分布
for name in ("问题", "未复现", "精度", "使用"):
    ws = wb[name]
    hdr = [c.value for c in ws[1]]
    cats = Counter()
    frames = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cats[str(row[0]).strip()] += 1
        if len(row) > 3 and row[3] is not None:
            frames[str(row[3]).strip()] += 1
    print(f"\n== {name}: 类别取值({len(cats)}) ==")
    for k, v in cats.most_common(40):
        print(f"  {k}: {v}")
    print(f"  帧数取值: {dict(frames.most_common(10))}")
