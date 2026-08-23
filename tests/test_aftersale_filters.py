# -*- coding: utf-8 -*-
"""售后面板三个状态筛选下拉——数据层回归测试

覆盖记录与统计页筛选栏从单一「是否解决」还原为三个独立下拉
（是否解决 / 是否我们主动发起 / 是否是我们的问题）后新增的数据层：
- _build_where：is_initiative / is_our_problem 条件拼接（SQL 与参数）
- query_page：按 is_initiative / is_our_problem 过滤
- query_with_stats：列表跟随新筛选、stats 保持周期全景（与 resolved 口径一致）
- export_xlsx：新参数透传不回归

隔离方式同 test_aftersale_batch_ops：tmp SQLite + monkeypatch，
不触碰真实 tables.db；周期模式固定为周二起保证归属确定性。
"""
import sqlite3

import pytest

import database.backend as backend
import database.table_db as table_db
import database.aftersale_db as adb
from database import schema

TUE = {"type": "tue", "start": "", "span": 7}


@pytest.fixture
def db(monkeypatch, tmp_path):
    """临时库：建 aftersale_records 表 + 固定周期模式，隔离 settings.json"""
    path = str(tmp_path / "t.db")
    monkeypatch.setattr(table_db, "DB_PATH", path)
    monkeypatch.setattr(table_db, "_conn", None)
    monkeypatch.setattr(table_db, "_ensure_initialized", lambda c: None)
    monkeypatch.setattr(backend, "is_mysql_test_mode", lambda: False)
    monkeypatch.setattr(adb, "load_cycle_mode", lambda: dict(TUE))
    sl = sqlite3.connect(path)
    sl.executescript(schema.to_sqlite_ddl("aftersale_records"))
    sl.commit()
    sl.close()
    return path


def _seed(db, recs):
    """插入记录列表；每条为 (is_initiative, is_our_problem, resolved, idx)"""
    sl = sqlite3.connect(db)
    for i, (initiative, our_problem, resolved) in enumerate(recs):
        sl.execute(
            "INSERT INTO aftersale_records "
            "(created_at, occurred_at, creator, issue_type, table_no, "
            "room_name, region, problem, cause, resolved, is_initiative, "
            "is_our_problem, solution, resolver, response_time, snk_code, "
            "device_code, cycle_start) "
            "VALUES (?, ?, 'tester', '硬件问题', ?, ?, '', '问题X', '', "
            "?, ?, ?, '', '', '', '', '', '')",
            ("2026-08-20 10:00:00", "2026-08-20", f"T{i}", f"room{i}",
             resolved, initiative, our_problem))
    sl.commit()
    ids = [r[0] for r in sl.execute(
        "SELECT id FROM aftersale_records ORDER BY id").fetchall()]
    sl.close()
    return ids


# ==================== _build_where 条件拼接 ====================

def test_build_where_empty():
    where, params = adb._build_where("", "", "")
    assert where == ""
    assert params == []


def test_build_where_only_initiative():
    where, params = adb._build_where("", "", "", is_initiative="是")
    assert "is_initiative = ?" in where
    assert params == ["是"]
    assert "is_our_problem" not in where


def test_build_where_only_our_problem():
    where, params = adb._build_where("", "", "", is_our_problem="否")
    assert "is_our_problem = ?" in where
    assert params == ["否"]
    assert "is_initiative" not in where


def test_build_where_combined():
    where, params = adb._build_where(
        "球房", "硬件问题", "否", is_initiative="是", is_our_problem="否")
    for frag in ("issue_type = ?", "resolved = ?",
                 "is_initiative = ?", "is_our_problem = ?"):
        assert frag in where
    assert "LIKE ?" in where  # 关键词模糊匹配仍存在
    assert params[:4] == ["硬件问题", "否", "是", "否"]
    assert len(params) == 4 + len(adb._SEARCH_FIELDS)


# ==================== query_page 过滤 ====================

def test_query_page_filter_initiative(db):
    _seed(db, [
        ("是", "是", "是"),
        ("否", "是", "否"),
        ("是", "否", "否"),
    ])
    total, rows = adb.query_page(1, 50, is_initiative="是")
    assert total == 2
    assert all(r["is_initiative"] == "是" for r in rows)


def test_query_page_filter_our_problem(db):
    _seed(db, [
        ("是", "是", "是"),
        ("否", "是", "否"),
        ("是", "否", "否"),
    ])
    total, rows = adb.query_page(1, 50, is_our_problem="否")
    assert total == 1
    assert rows[0]["is_our_problem"] == "否"


def test_query_page_filter_combined(db):
    _seed(db, [
        ("是", "是", "是"),
        ("是", "否", "否"),
        ("否", "否", "否"),
        ("否", "是", "是"),
    ])
    total, rows = adb.query_page(1, 50, is_initiative="是",
                                 is_our_problem="否", resolved="否")
    assert total == 1
    r = rows[0]
    assert r["is_initiative"] == "是"
    assert r["is_our_problem"] == "否"
    assert r["resolved"] == "否"


def test_query_page_no_filters_returns_all(db):
    _seed(db, [
        ("是", "是", "是"),
        ("否", "否", "否"),
    ])
    total, rows = adb.query_page(1, 50)
    assert total == 2 and len(rows) == 2


# ==================== query_with_stats 口径 ====================

def test_query_with_stats_list_filtered_stats_full(db):
    """新筛选只过滤列表，stats 仍为周期全景（与 resolved 口径一致）"""
    _seed(db, [
        ("是", "是", "是"),
        ("否", "是", "否"),
        ("是", "否", "否"),
        ("否", "否", "是"),
    ])
    total, rows, stats = adb.query_with_stats(1, 50, is_initiative="是")
    assert total == 2 and len(rows) == 2
    assert all(r["is_initiative"] == "是" for r in rows)
    assert stats["total"] == 4           # 全景
    assert stats["resolved"] == 2
    assert stats["initiative"] == 2
    # 与不传新筛选时的 stats 完全一致
    _t2, _r2, stats_plain = adb.query_with_stats(1, 50)
    assert stats == stats_plain


def test_query_with_stats_our_problem_filter(db):
    _seed(db, [
        ("是", "是", "是"),
        ("否", "否", "否"),
        ("是", "否", "否"),
    ])
    total, rows, stats = adb.query_with_stats(1, 50, is_our_problem="否")
    assert total == 2 and len(rows) == 2
    assert all(r["is_our_problem"] == "否" for r in rows)
    assert stats["total"] == 3
    assert stats["resolved"] == 1


# ==================== 自然月模式（type=month 配置） ====================

MONTH = {"type": "month", "start": "", "span": 7}


def _seed_months(db):
    """插入跨月记录：2026-07 两条 + 2026-08 一条（tue 模式下各属不同周周期）"""
    sl = sqlite3.connect(db)
    for i, (occ, initiative, our_problem, resolved) in enumerate([
            ("2026-07-05", "是", "是", "是"),
            ("2026-07-20", "否", "否", "否"),
            ("2026-08-10", "是", "否", "是")]):
        sl.execute(
            "INSERT INTO aftersale_records "
            "(created_at, occurred_at, creator, issue_type, table_no, "
            "room_name, region, problem, cause, resolved, is_initiative, "
            "is_our_problem, solution, resolver, response_time, snk_code, "
            "device_code, cycle_start) "
            "VALUES (?, ?, 'tester', '硬件问题', ?, ?, '', '问题X', '', "
            "?, ?, ?, '', '', '', '', '', '')",
            (occ + " 10:00:00", occ, f"T{i}", f"room{i}",
             resolved, initiative, our_problem))
    sl.commit()
    sl.close()


def test_query_with_stats_month_mode(db, monkeypatch):
    """自然月模式：列表与统计均按自然月归属过滤（type=month 配置生效）"""
    monkeypatch.setattr(adb, "load_cycle_mode", lambda: dict(MONTH))
    _seed_months(db)
    # 2026-07：两条
    total, rows, stats = adb.query_with_stats(
        1, 50, cycle_start="2026/07/01")
    assert total == 2 and len(rows) == 2
    assert stats["total"] == 2
    assert stats["resolved"] == 1
    # 2026-08：一条
    total, rows, stats = adb.query_with_stats(
        1, 50, cycle_start="2026/08/01")
    assert total == 1 and len(rows) == 1
    assert stats["total"] == 1


def test_query_page_month_mode_isolated_from_week(db):
    """口径隔离：周模式查询不命中 month 的 cycle_start（防止混用）"""
    _seed_months(db)
    total, rows, _ = adb.query_with_stats(
        1, 50, cycle_start="2026/07/01")
    assert total == 0  # tue 口径下 7/01 不是周期起点


def test_get_cycle_options_month(db, monkeypatch):
    """自然月模式选项：返回有数据的自然月（每月 1 号，降序）"""
    monkeypatch.setattr(adb, "load_cycle_mode", lambda: dict(MONTH))
    _seed_months(db)
    assert adb.get_cycle_options() == ["2026/08/01", "2026/07/01"]
    # 周口径不受影响：tue 模式各记录归属 06/30、07/14、08/04 周期
    monkeypatch.setattr(adb, "load_cycle_mode", lambda: dict(TUE))
    assert adb.get_cycle_options() == ["2026/08/04", "2026/07/14",
                                       "2026/06/30"]


def test_export_xlsx_month_mode(db, tmp_path, monkeypatch):
    """自然月导出：周期筛选按自然月，周期列显示月份标签"""
    pytest.importorskip("openpyxl")
    monkeypatch.setattr(adb, "load_cycle_mode", lambda: dict(MONTH))
    _seed_months(db)
    out = str(tmp_path / "out_month.xlsx")
    n = adb.export_xlsx(out, cycle_start="2026/07/01")
    assert n == 2
    from openpyxl import load_workbook
    wb = load_workbook(out, data_only=True)
    ws = wb["售后记录"]
    labels = [ws.cell(row=r, column=14).value for r in (2, 3)]
    assert labels == ["2026-07", "2026-07"]  # 周期列显示月份标签


# ==================== export_xlsx 透传 ====================

def test_export_xlsx_filters_passthrough(db, tmp_path):
    """新参数透传：导出条数跟随筛选口径，不回归旧签名

    venv 测试环境未装 openpyxl（基线 145 无导出用例），缺依赖时跳过；
    有 openpyxl 的环境（GUI 运行环境）完整验证透传与落盘。
    """
    pytest.importorskip("openpyxl")
    _seed(db, [
        ("是", "是", "是"),
        ("否", "是", "否"),
        ("是", "否", "否"),
        ("否", "否", "是"),
    ])
    out = str(tmp_path / "out.xlsx")
    n_all = adb.export_xlsx(out)
    assert n_all == 4
    n_init = adb.export_xlsx(out, is_initiative="是")
    assert n_init == 2
    n_ours = adb.export_xlsx(out, is_our_problem="否")
    assert n_ours == 2
    n_both = adb.export_xlsx(out, is_initiative="是", is_our_problem="是")
    assert n_both == 1
