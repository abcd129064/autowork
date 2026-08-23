# -*- coding: utf-8 -*-
"""临时：解析 在线模板.xlsx 全部 sheet 结构与样例数据（验证后删除）"""
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\shen_zhe\Desktop\在线模板.xlsx", data_only=True)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("=" * 60)
    print(f"sheet={name}  max_row={ws.max_row}  max_col={ws.max_column}  dims={ws.dimensions}")
    # 表头
    hdr = [c.value for c in ws[1]]
    print("表头:", hdr)
    # 合并单元格
    print("merged:", list(ws.merged_cells.ranges)[:20])
    # 前 5 行数据
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(ws.max_row, 6), values_only=True), 2):
        print(f"  行{i}:", row)
    # 列宽
    widths = {k: round(v.width, 1) if v.width else None for k, v in ws.column_dimensions.items()}
    print("列宽:", widths)
