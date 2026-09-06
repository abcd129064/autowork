# -*- coding: utf-8 -*-
"""售后记录数据层（SQLite / MySQL 双后端，自动跟随 MySQL 测试开关）

职责：
- insert_record / update_record / delete_record：售后记录增删改
- query_page / query_with_stats：筛选 + 分页 + 统计（周期/类型/状态/主动发起/我方问题/关键词）
- get_cycle_options：周期下拉选项（库中记录实际归属的周期）
- get_field_candidates：问题/解决人/地区动态候选
- export_xlsx：按筛选条件导出（表头与 售后问题汇总8月.xlsx 对齐）
- import_excel_rows：一次性导入历史 Excel

连接复用 table_db 双后端路由（SQLite 单连接 / MySQL thread-local），
MySQL 模式下多人各自提交即提交即落库，其他用户刷新/手动同步后可见。
周期规则：周二开始、周一结束为默认；可在售后面板「周期设置」切换为自然周
（周一~周日）或自定义（起始日+周期天数）。
周期归属统一按记录发生时间 occurred_at 动态计算（缺失时回退填写时间
created_at）——列表、统计、周期下拉、导出共用 cycle_start_of 同一规则。
性能（2026-09-06 S3）：周期筛选改为走落库的 cycle_start 物化列（等值过滤，
idx_aftersale_cycle 索引可用，旧表达式过滤索引失效全表扫），该列由
insert_record 落库时维护 + 周期配置保存/面板首次加载时 recalc_cycle_starts
幂等重算保证与动态口径一致；存量空值行由 _ensure_cycle_materialized 兜底回填。
"""

import os
from datetime import datetime, timedelta

from database import table_db

# ==================== 字段枚举 ====================

# 类型
ISSUE_TYPES = (
    "硬件问题", "程序相关", "识别问题", "直播相关", "操作问题",
    "其他问题", "相机偏移", "新球助手", "安装调试", "不能扫码", "待查",
)

# 地区预置
REGIONS_PRESET = ("上海", "云南", "四川", "广东", "新疆", "江苏", "江西", "湖南", "西藏")

# 响应时间预置档位
RESPONSE_TIME_PRESET = ("1分钟内", "5分钟内", "30分钟内", "1小时内", "1小时以上")

# 团队默认人员
_CREW_PRESET = ("张峻涛", "沈喆", "孙跃源", "吴斌", "贺勤")

# 记录字段
RECORD_FIELDS = (
    "created_at", "occurred_at", "creator", "issue_type", "table_no",
    "room_name", "region", "problem", "cause", "resolved",
    "is_initiative", "is_our_problem", "solution", "resolver",
    "response_time", "is_important", "snk_code", "device_code", "cycle_start",
    "updated_at",
)

# 售后业务键：SQLite 与 MySQL 两侧 id 各自
# 增长会撞车，按 created_at, creator, table_no, problem 判定同一条记录。
# merge_back / 历史 mysql_sync 推送共用此定义。
RECORD_KEY_COLS = ("created_at", "creator", "table_no", "problem")

# 关键词搜索覆盖列，全字段模糊匹配
_SEARCH_FIELDS = (
    "table_no", "room_name", "problem", "region",
    "cause", "solution", "resolver", "creator",
)

# 导出表头
_EXPORT_HEADERS = (
    ("issue_type", "类型"), ("room_name", "球房"), ("table_no", "桌号"),
    ("region", "地区"), ("problem", "问题"), ("cause", "发生原因"),
    ("resolved", "是否解决"), ("solution", "解决方案"), ("resolver", "解决人"),
    ("response_time", "响应时间"), ("created_at", "填写时间"),
    ("occurred_at", "发生时间"), ("creator", "填写人"),
    ("cycle_start", "周期"),
)


# ==================== 周期计算 ====================

# 周期模式：tue=周二起 / mon=自然周 / custom=自定义起始日+周期天数 / month=自然月
CYCLE_MODE_DEFAULT = {"type": "tue", "start": "", "span": 7}
_CYCLE_KEY = "aftersale_cycle"
_VALID_TYPES = ("tue", "mon", "custom", "month")

# S1（2026-09-06）：周期模式进程内缓存——get_cycle_options 会对每个 DISTINCT
# 日期调 cycle_start_of → load_cycle_mode，旧实现每次 open+json.load 读盘，
# 241 个日期即 241 次读盘（占周期下拉耗时 74%）。save_cycle_mode 写盘成功后
# 置空失效（与 backend.py 的 MySQL settings 缓存同策略）；测试均 monkeypatch
# 注入替换函数，天然绕过缓存，不受影响。
_cycle_mode_cache: dict | None = None


def load_cycle_mode() -> dict:
    """读取周期模式设置，settings.json 的 aftersale_cycle，缺省/非法回退周二起

    命中进程内缓存直接返回副本（防止调用方污染缓存）。
    """
    global _cycle_mode_cache
    if _cycle_mode_cache is not None:
        return dict(_cycle_mode_cache)
    import json
    from core.app_paths import get_app_dir

    cfg = dict(CYCLE_MODE_DEFAULT)
    try:
        path = os.path.join(get_app_dir(), "settings.json")
        val = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                val = json.load(f).get(_CYCLE_KEY) or {}
        if isinstance(val, dict) and val.get("type") in _VALID_TYPES:
            cfg["type"] = val["type"]
        if isinstance(val, dict) and val.get("start"):
            cfg["start"] = str(val["start"])
        if isinstance(val, dict):
            try:
                raw = val.get("span")
                cfg["span"] = max(1, int(raw if raw not in (None, "") else 7))
            except (TypeError, ValueError):
                cfg["span"] = 7
    except Exception:
        pass
    _cycle_mode_cache = dict(cfg)
    return cfg


def save_cycle_mode(mode: dict) -> dict:
    """保存周期模式设置，合并写，保留 settings.json 其余字段，返回规范化配置"""
    global _cycle_mode_cache
    import json
    from core.app_paths import get_app_dir

    cfg = load_cycle_mode()
    if mode.get("type") in _VALID_TYPES:
        cfg["type"] = mode["type"]
    if mode.get("start"):
        cfg["start"] = str(mode["start"])
    try:
        raw = mode.get("span")
        cfg["span"] = max(1, int(raw if raw not in (None, "") else 7))
    except (TypeError, ValueError):
        cfg["span"] = 7
    path = os.path.join(get_app_dir(), "settings.json")
    try:
        data = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        data[_CYCLE_KEY] = cfg
        # S3 返工：置重算待办标志——周期配置已切换而物化列尚未重算，
        # recalc_cycle_starts 成功后清除（失败残留，由面板首载自愈追平）
        data[_CYCLE_RECALC_PENDING_KEY] = True
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        # 写盘成功：置空周期模式缓存，下次读取按新配置生效
        _cycle_mode_cache = None
    except Exception:
        pass
    return cfg


# ==================== 周期物化列重算待办标志 ====================

# S3 返工（2026-09-06）：保存周期成功 → 重算 worker 失败（DB 瞬时锁等）
# 的场景下，周期下拉按新口径（动态计算）而数据按旧口径（物化列）过滤，
# 命中错/空且无自动追平路径。本标志实现自愈：save 置 True → recalc
# 成功清除 → 面板每次打开时检查，有残留即全量重算（无需用户干预）。
_CYCLE_RECALC_PENDING_KEY = "aftersale_cycle_recalc_pending"


def _load_recalc_pending() -> bool:
    """读取重算待办标志（settings.json 布尔键，缺省/异常回退 False）"""
    import json
    from core.app_paths import get_app_dir
    try:
        path = os.path.join(get_app_dir(), "settings.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return bool(json.load(f).get(_CYCLE_RECALC_PENDING_KEY))
    except Exception:
        pass
    return False


def _save_recalc_pending(flag: bool) -> None:
    """合并写重算待办标志（False 时移除键，稳态 settings.json 无冗余键）"""
    import json
    from core.app_paths import get_app_dir
    path = os.path.join(get_app_dir(), "settings.json")
    data = {}
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    if flag:
        data[_CYCLE_RECALC_PENDING_KEY] = True
    else:
        data.pop(_CYCLE_RECALC_PENDING_KEY, None)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ==================== 常用句 ====================

_QPC_KEY = "aftersale_quick_phrases"
# 预置常见问题描述
_QPC_DEFAULT = [
    "主机没有开机", "遥控器没反应", "程序没了", "不能扫码",
    "识别不了", "记分牌显示不出来",
]


def load_quick_phrases() -> list:
    """读取常用句列表，settings.json 持久化"""
    import json
    from core.app_paths import get_app_dir

    try:
        path = os.path.join(get_app_dir(), "settings.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                v = json.load(f).get(_QPC_KEY)
            if isinstance(v, list):
                return [str(x) for x in v if str(x).strip()]
    except Exception:
        pass
    return list(_QPC_DEFAULT)


def add_quick_phrase(text: str) -> list:
    """新增常用句，写回 settings.json，返回更新后的列表"""
    import json
    from core.app_paths import get_app_dir

    phrases = load_quick_phrases()
    t = str(text or "").strip()
    if t and t not in phrases:
        phrases.insert(0, t)
    path = os.path.join(get_app_dir(), "settings.json")
    try:
        data = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        data[_QPC_KEY] = phrases
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return phrases


def cycle_span_days() -> int:
    """当前模式周期天数：tue/mon 固定 7 天，custom 取配置值（≥1），month 取当月天数"""
    import calendar

    mode = load_cycle_mode()
    if mode["type"] == "month":
        now = datetime.now()
        return calendar.monthrange(now.year, now.month)[1]
    if mode["type"] == "custom":
        try:
            return max(1, int(mode.get("span") or 7))
        except (TypeError, ValueError):
            return 7
    return 7


def cycle_start_of(dt: datetime) -> str:
    """计算给定时间所属周期的起始日，格式 yyyy/MM/dd

    模式：
    - tue（默认）：周二开始周一结束，days_since_tue=(weekday-1)%7
    - mon（自然周）：周一开始周日结束，days_since_mon=weekday
    - custom：从配置起始日起每 span 天一个周期，历史周期向过去对齐
    - month：自然月，当月 1 号为起始日
    """
    mode = load_cycle_mode()
    if mode.get("type") == "month":
        return dt.replace(day=1).strftime("%Y/%m/%d")
    if mode["type"] == "mon":
        start = dt - timedelta(days=dt.weekday())
    elif mode["type"] == "custom":
        start0 = _parse_occurred(mode.get("start")) or dt
        span = max(1, int(mode.get("span") or 7))
        diff = (dt.date() - start0.date()).days
        start = start0 + timedelta(days=(diff // span) * span)
    else:  # tue 默认
        start = dt - timedelta(days=(dt.weekday() - 1) % 7)
    return start.strftime("%Y/%m/%d")


def current_cycle_start() -> str:
    """当前周期起始日"""
    return cycle_start_of(datetime.now())


def _parse_occurred(occurred: str):
    """解析日期串 YYYY-MM-DD，非法返回 None 由调用方回退"""
    try:
        return datetime.strptime(str(occurred or "").strip()[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def cycle_label(cycle_start: str) -> str:
    """周期展示标签：周模式 '08/19 - 08/25'（起始日 + 周期天数-1，custom 按配置 span）；month 模式 '2026-08'（自然月）"""
    try:
        start = datetime.strptime(str(cycle_start).strip(), "%Y/%m/%d")
    except (ValueError, TypeError):
        return str(cycle_start or "")
    if load_cycle_mode().get("type") == "month":
        return f"{start:%Y}-{start:%m}"
    end = start + timedelta(days=cycle_span_days() - 1)
    return f"{start:%m/%d} - {end:%m/%d}"


# ==================== 周期归属 ====================

def cycle_date_range(cycle_start: str) -> tuple:
    """周期起止日期（date, date）：month 模式为整月，其余为起始日 + span-1 天

    供统计弹窗「当前周期」趋势范围使用；month 模式按周期自身所在月份计算
    （不依赖当前时刻，历史周期同样正确）。非法输入返回 (None, None)。
    """
    start = _parse_occurred(cycle_start)
    if start is None:
        # 周期筛选/落库格式为 yyyy/MM/dd
        try:
            start = datetime.strptime(
                str(cycle_start or "").strip()[:10], "%Y/%m/%d")
        except (ValueError, TypeError):
            return None, None
    s = start.date()
    if load_cycle_mode().get("type") == "month":
        import calendar
        e = s.replace(day=calendar.monthrange(s.year, s.month)[1])
    else:
        e = s + timedelta(days=cycle_span_days() - 1)
    return s, e


def _record_cycle(occurred_at: str, created_at: str):
    """记录归属周期：优先按 occurred_at 计算，缺失/非法时回退 created_at，

    两者均无返回 None。列表筛选、统计、下拉选项、导出全部经此函数归属，
    保证四处口径一致；不依赖冗余落库的 cycle_start 字段。
    归属按当前周期配置（含 month 自然月）。
    """
    for raw in (occurred_at, created_at):
        dt = _parse_occurred(raw)
        if dt:
            return cycle_start_of(dt)
    return None


def _match_cycle(record: dict, cycle_start: str) -> bool:
    """记录是否属于所选周期：按记录时间动态计算后匹配，空周期不过滤"""
    if not cycle_start:
        return True
    return _record_cycle(record.get("occurred_at"),
                         record.get("created_at")) == str(cycle_start).strip()


# ==================== 连接与工具 ====================

def _conn():
    """复用 table_db 双后端连接，SQLite 单连接 和 MySQL thread-local"""
    return table_db.get_conn()


def _lookup_table_binding(table_no: str) -> tuple:
    """按桌号精确匹配球桌管理库，返回 snk_code, device_code

    桌号自由文本允许非标格式，匹配不到返回空串不阻断。device_code 取该球桌最近一期 kd 状态记录。
    """
    name = str(table_no or "").strip()
    if not name:
        return "", ""
    snk = table_db.get_snk_by_name(name)
    device = ""
    try:
        conn = _conn()
        row = conn.execute(
            "SELECT device_code FROM kd_status WHERE TRIM(table_id) = ? "
            "ORDER BY file_path DESC LIMIT 1", (name,)).fetchone()
        device = str(row[0] or "").strip() if row else ""
    except Exception:
        pass
    return str(snk or ""), device


# ==================== 增删改 ====================

def insert_record(record: dict) -> int:
    """新增售后记录，返回新记录 id

    created_at 自动取填写时刻；occurred_at（发生时间）缺省取当日；
    cycle_start 按发生时间归属周期，发生时间缺失时回退填写时间；
    snk_code/device_code 未提供时按桌号精确匹配球桌管理库自动带出。
    """
    now = datetime.now()
    created_at = now.strftime("%Y-%m-%d %H:%M:%S")
    occurred = str(record.get("occurred_at") or "").strip() or created_at[:10]
    occ_dt = _parse_occurred(occurred)
    cycle = cycle_start_of(occ_dt) if occ_dt else cycle_start_of(now)
    snk = str(record.get("snk_code") or "").strip()
    device = str(record.get("device_code") or "").strip()
    if not snk:
        snk, device = _lookup_table_binding(record.get("table_no"))
    conn = _conn()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.execute(
        "INSERT INTO aftersale_records "
        "(created_at, occurred_at, creator, issue_type, table_no, room_name, "
        "region, problem, cause, resolved, is_initiative, is_our_problem, "
        "solution, resolver, response_time, snk_code, device_code, "
        "cycle_start, updated_at, is_important) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (created_at, occurred,
         str(record.get("creator") or ""),
         str(record.get("issue_type") or ""),
         str(record.get("table_no") or ""),
         str(record.get("room_name") or ""),
         str(record.get("region") or ""),
         str(record.get("problem") or ""),
         str(record.get("cause") or ""),
         str(record.get("resolved") or "否"),
         str(record.get("is_initiative") or "否"),
         str(record.get("is_our_problem") or "是"),
         str(record.get("solution") or ""),
         str(record.get("resolver") or ""),
         str(record.get("response_time") or ""),
         snk, device, cycle, now_str,
         int(bool(record.get("is_important"))) ))
    conn.commit()
    _invalidate_field_cands_cache()
    return cur.lastrowid


def update_record(record: dict) -> int:
    """按 id 更新记录（created_at/cycle_start 保留原值），返回受影响行数

    多人场景：MySQL autocommit 提交后其他用户刷新即可见。
    """
    rec_id = record.get("id")
    if not rec_id:
        return 0
    # 编辑不改动填写时间；cycle_start 若缺失优先按发生时间重算，
    # 发生时间也缺失时回退填写时间，保证改发生时间后周期跟随
    occurred = str(record.get("occurred_at") or "").strip()
    created_at = str(record.get("created_at") or "")
    cycle = str(record.get("cycle_start") or "")
    if not cycle:
        occ_dt = _parse_occurred(occurred)
        if occ_dt:
            cycle = cycle_start_of(occ_dt)
        elif created_at:
            try:
                cycle = cycle_start_of(
                    datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S"))
            except ValueError:
                cycle = ""
    snk = str(record.get("snk_code") or "").strip()
    device = str(record.get("device_code") or "").strip()
    if not snk:
        snk, device = _lookup_table_binding(record.get("table_no"))
    conn = _conn()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.execute(
        "UPDATE aftersale_records SET occurred_at=?, creator=?, issue_type=?, "
        "table_no=?, room_name=?, region=?, problem=?, cause=?, resolved=?, "
        "is_initiative=?, is_our_problem=?, solution=?, "
        "resolver=?, response_time=?, snk_code=?, device_code=?, cycle_start=?, "
        "updated_at=?, is_important=? "
        "WHERE id=?",
        (occurred,
         str(record.get("creator") or ""),
         str(record.get("issue_type") or ""),
         str(record.get("table_no") or ""),
         str(record.get("room_name") or ""),
         str(record.get("region") or ""),
         str(record.get("problem") or ""),
         str(record.get("cause") or ""),
         str(record.get("resolved") or "否"),
         str(record.get("is_initiative") or "否"),
         str(record.get("is_our_problem") or "是"),
         str(record.get("solution") or ""),
         str(record.get("resolver") or ""),
         str(record.get("response_time") or ""),
         snk, device, cycle, now_str,
         int(bool(record.get("is_important"))), rec_id))
    conn.commit()
    _invalidate_field_cands_cache()
    return cur.rowcount


def delete_record(rec_id) -> int:
    """按 id 删除记录，返回受影响行数"""
    if not rec_id:
        return 0
    conn = _conn()
    cur = conn.execute("DELETE FROM aftersale_records WHERE id = ?", (rec_id,))
    conn.commit()
    _invalidate_field_cands_cache()
    return cur.rowcount


def mark_resolved_batch(rec_ids) -> int:
    """批量标记已解决，返回受影响行数

    供列表「一键标记已解决 / 批量标记」使用：区别于 update_record 的全字段
    回写，不依赖 UI 提供其余字段原值，不会误改其他字段。
    """
    ids = [i for i in rec_ids if i]
    if not ids:
        return 0
    conn = _conn()
    qs = ", ".join(["?"] * len(ids))
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.execute(
        f"UPDATE aftersale_records SET resolved = '是', updated_at = ? "
        f"WHERE id IN ({qs})", [now_str] + list(ids))
    conn.commit()
    _invalidate_field_cands_cache()
    return cur.rowcount


def delete_records(rec_ids) -> int:
    """按 id 批量删除记录，返回受影响行数"""
    ids = [i for i in rec_ids if i]
    if not ids:
        return 0
    conn = _conn()
    qs = ", ".join(["?"] * len(ids))
    cur = conn.execute(f"DELETE FROM aftersale_records WHERE id IN ({qs})", ids)
    conn.commit()
    _invalidate_field_cands_cache()
    return cur.rowcount


# ==================== 查询 ====================

# 记录日期表达式：优先 occurred_at（格式合法时），否则回退 created_at，
# 与 Python _record_cycle 的回退语义一致（LIKE '____-__-__%' 校验
# yyyy-MM-dd 外形；空串/非法格式均回退）。两侧方言语法一致。
_RECORD_DATE_EXPR = (
    "substr(CASE WHEN occurred_at LIKE '____-__-__%' "
    "THEN occurred_at ELSE created_at END, 1, 10)")


# 非法周期起点判定：传入的 cycle_start 不是当前模式下的真实周期起点
# （如 tue 模式传入周三日期），语义上无任何记录命中 → 追加 WHERE 1=0。
def _cycle_is_valid_start(cycle_start: str):
    """周期起点合法性判定，返回 True / False / None

    - True：当前模式下的真实周期起点（yyyy/MM/dd，cycle_start_of 校验
      起点回到自身），正常走物化列过滤；
    - False：格式合法但非本模式真实周期起点（周/月口径混用）→ 0 命中；
    - None：无法解析（与旧实现一致，调用方不追加过滤）。
    """
    cyc = str(cycle_start or "").strip()
    try:
        start = datetime.strptime(cyc, "%Y/%m/%d")
    except (ValueError, TypeError):
        return None
    return cycle_start_of(start) == cyc


# 物化列兜底回填的连续失效计数（日期形似但不可解析/回填异常时防循环）
_cycle_recalc_futile = 0


def _legacy_cycle_rows_pending() -> bool:
    """是否存在 cycle_start 为空但记录日期可解析的存量行（物化列待回填）

    走 idx_aftersale_cycle 索引探测（cycle_start = '' 等值），干净库上
    微秒级。LIKE '____-__-__%' 与 _parse_occurred 的 yyyy-MM-dd 外形校验
    一致，仅命中可能被回填的行。
    """
    conn = _conn()
    row = conn.execute(
        "SELECT 1 FROM aftersale_records WHERE cycle_start = '' AND "
        "(occurred_at LIKE '____-__-__%' OR created_at LIKE '____-__-__%') "
        "LIMIT 1").fetchone()
    return row is not None


def _ensure_cycle_materialized() -> None:
    """周期物化列兜底：检测到存量空值行时幂等回填一次（S3）

    周期筛选走 cycle_start 物化列（索引化），直写库的历史数据（早期
    Excel 导入/外部写入）该列可能为空，物化过滤会漏掉它们——检测到
    即同步重算（_record_cycle 同口径，幂等）。回填后仍有残留（日期
    形似但不可解析，_record_cycle 返回 None 写 ''）或回填异常（只读
    库等）时递增失效计数，进程内最多重试 2 次，避免每次查询全表重算。
    """
    global _cycle_recalc_futile
    if _cycle_recalc_futile >= 2:
        return
    try:
        if _legacy_cycle_rows_pending():
            recalc_cycle_starts()
            if _legacy_cycle_rows_pending():
                _cycle_recalc_futile += 1
            else:
                _cycle_recalc_futile = 0
        else:
            _cycle_recalc_futile = 0
    except Exception:
        _cycle_recalc_futile += 1


def recalc_cycle_starts(batch: int = 2000) -> int:
    """按当前周期配置全表重算 cycle_start 物化列，返回更新行数（幂等）

    分批（默认 2000 行/批）SELECT id, occurred_at, created_at,
    cycle_start → _record_cycle 计算新周期起点 → 与现值不同的收集后
    executemany UPDATE；_record_cycle 返回 None（两日期均无法解析）的
    记录写 ''。周期配置保存后、面板首次加载兜底、存量数据一次性回填
    共用本函数；二次执行无差异时返回 0（无感知）。成功路径清除
    重算待办标志（先读后写，稳态无标志时不产生 settings.json 写盘）。
    """
    conn = _conn()
    updated = 0
    last_id = 0
    while True:
        rows = conn.execute(
            "SELECT id, occurred_at, created_at, cycle_start "
            "FROM aftersale_records WHERE id > ? ORDER BY id LIMIT ?",
            (last_id, int(batch))).fetchall()
        if not rows:
            break
        last_id = rows[-1][0]
        updates = []
        for rec_id, occ, cre, cur in rows:
            new = _record_cycle(occ, cre) or ""
            if str(cur or "") != new:
                updates.append((new, rec_id))
        if updates:
            conn.executemany(
                "UPDATE aftersale_records SET cycle_start = ? WHERE id = ?",
                updates)
            updated += len(updates)
    conn.commit()
    # 重算成功 → 清除待办标志（任何触发路径统一在此闭环；失败/异常不清除，
    # 由面板下次打开经 recalc_cycle_starts_on_load 自愈）
    try:
        if _load_recalc_pending():
            _save_recalc_pending(False)
    except Exception:
        pass
    return updated


def recalc_cycle_starts_on_load() -> int:
    """面板首载兜底：有待办标志或存量空值行 → 全量重算，否则仅索引探测

    S3 返工（2026-09-06）：上次保存周期后的重算失败会残留
    aftersale_cycle_recalc_pending 标志——周期下拉按新口径（动态计算）、
    数据按旧口径（物化列）过滤，命中错/空。本函数在面板每次打开时检查：
    有待办标志（无论空值探测结果）或探测到空值存量行 → 全量
    recalc_cycle_starts（成功即清标志），否则走微秒级空值探测返回 0，
    正常路径零额外成本。幂等，可放心重复调用。
    """
    if _load_recalc_pending() or _legacy_cycle_rows_pending():
        return recalc_cycle_starts()
    return 0


def _append_cycle_where(where: str, params: list, cycle_start: str) -> tuple:
    """把周期筛选并入 WHERE，返回 (where, params)

    S3（2026-09-06）：由 _RECORD_DATE_EXPR 表达式范围过滤（索引失效，
    EXPLAIN=SCAN，每次加载 3 条全表扫）改为 cycle_start 物化列等值过滤，
    idx_aftersale_cycle 两端索引直接可用。传入非当前模式真实周期起点
    时追加 1=0（与旧实现「非起点返回 0 条」等价，防止周/月口径混用），
    无法解析的串不追加过滤（与旧实现一致）；物化列由 insert_record
    落库维护 + 周期保存/首次加载重算保证一致，存量空值行先经
    _ensure_cycle_materialized 幂等回填，旧数据不漏。
    """
    cyc = str(cycle_start or "").strip()
    if not cyc:
        return where, params
    valid = _cycle_is_valid_start(cyc)
    if valid is None:
        return where, params
    if not valid:
        # 非本模式真实周期起点 → 无记录命中（与旧实现 total=0 等价）
        where += (" AND " if where else " WHERE ") + " 1 = 0"
        return where, params
    _ensure_cycle_materialized()
    where += (" AND " if where else " WHERE ") + "cycle_start = ?"
    return where, params + [cyc]


def _build_where(keyword: str, issue_type: str, resolved: str,
                 is_initiative: str = "", is_our_problem: str = "") -> tuple:
    """构造筛选 WHERE 子句与参数（类型/状态/关键词/主动发起/我方问题）

    周期筛选不在本函数处理：由 _append_cycle_where 以 cycle_start 物化列
    等值过滤接入（与 _record_cycle 动态归属等价，由落库维护 + 重算保证
    一致），保证分页 COUNT/LIMIT 在数据库侧完成。
    统计口径说明：resolved / is_initiative / is_our_problem 为空才参与筛选；
    统计函数单独调用时传空串即得「已解决/未解决」分组基数与全景计数。
    """
    conds, params = [], []
    if issue_type:
        conds.append("issue_type = ?")
        params.append(str(issue_type).strip())
    if resolved:
        conds.append("resolved = ?")
        params.append(str(resolved).strip())
    if is_initiative:
        conds.append("is_initiative = ?")
        params.append(str(is_initiative).strip())
    if is_our_problem:
        conds.append("is_our_problem = ?")
        params.append(str(is_our_problem).strip())
    kw = str(keyword or "").strip()
    if kw:
        like = f"%{kw}%"
        kw_cond = " OR ".join([f"{f} LIKE ?" for f in _SEARCH_FIELDS])
        conds.append(f"({kw_cond})")
        params.extend([like] * len(_SEARCH_FIELDS))
    where = (" WHERE " + " AND ".join(conds)) if conds else ""
    return where, params


def query_page(page_no: int, page_size: int, keyword: str = "",
               cycle_start: str = "", issue_type: str = "",
               resolved: str = "", is_initiative: str = "",
               is_our_problem: str = "") -> tuple:
    """分页查询售后记录，返回 (total, rows)

    周期筛选 SQL 化（cycle_start 物化列等值过滤走索引，与 _record_cycle
    动态归属等价，由落库维护 + 重算保证一致）；COUNT + LIMIT/OFFSET
    数据库侧分页，避免数据量增大后全量取回（压测 10k→100k 全量取回
    p50 31.8→338.9ms，改造后与数据量无关，毫秒级）。
    """
    conn = _conn()
    where, params = _build_where(keyword, issue_type, resolved,
                                 is_initiative, is_our_problem)
    where, params = _append_cycle_where(where, params, cycle_start)
    total = conn.execute(
        f"SELECT COUNT(*) FROM aftersale_records{where}", params).fetchone()[0]
    offset = (max(1, page_no) - 1) * page_size
    cur = conn.execute(
        f"SELECT id, {', '.join(RECORD_FIELDS)} FROM aftersale_records"
        f"{where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [int(page_size), offset])
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return int(total or 0), rows


def query_with_stats(page_no: int, page_size: int, keyword: str = "",
                     cycle_start: str = "", issue_type: str = "",
                     resolved: str = "", is_initiative: str = "",
                     is_our_problem: str = "") -> tuple:
    """分页查询 + 同口径统计，返回 (total, rows, stats)

    周期按记录时间动态归属（与列表同规则），保证列表与统计一一对应；
    stats 统计不带 resolved/is_initiative/is_our_problem 筛选（否则
    已解决/未解决计数退化），返回 {total, resolved, unresolved,
    initiative, our_problem, rate}：initiative 为「我方主动发起」条数，
    our_problem 为「属我方问题」条数，rate 为已解决率（整数百分比）。
    """
    total, rows = query_page(page_no, page_size, keyword, cycle_start,
                             issue_type, resolved, is_initiative,
                             is_our_problem)
    conn = _conn()
    # 统计与列表同口径（keyword/issue_type/周期；不带 resolved/initiative/
    # our_problem 筛选）。单次 SQL 聚合替代旧实现的「二次全表取回 + Python
    # 逐条 sum」——压测 100k 下该路径 439.5ms → 毫秒级。
    where, params = _build_where(keyword, issue_type, "")
    where, params = _append_cycle_where(where, params, cycle_start)
    row = conn.execute(
        "SELECT COUNT(*), SUM(resolved = '是'), SUM(is_initiative = '是'), "
        "SUM(is_our_problem = '是') FROM aftersale_records" + where,
        params).fetchone()
    n_all = int(row[0] or 0)
    n_resolved = int(row[1] or 0)
    n_init = int(row[2] or 0)
    n_our = int(row[3] or 0)
    rate = int(round(n_resolved * 100 / n_all)) if n_all else 0
    stats = {"total": n_all, "resolved": n_resolved,
             "unresolved": n_all - n_resolved,
             "initiative": n_init, "our_problem": n_our, "rate": rate}
    return total, rows, stats


def query_stats_detail(keyword: str = "", cycle_start: str = "",
                       issue_type: str = "", trend_start: str = "",
                       trend_end: str = "") -> dict:
    """售后统计弹窗详细统计（与四卡片/列表完全同口径）

    与 query_with_stats 一致：统计不带 resolved/is_initiative/
    is_our_problem 筛选（否则已解决/未解决计数退化），仅按
    keyword + issue_type + cycle_start（按记录时间动态归属）过滤。
    返回：
    - summary: {total, resolved, unresolved, rate, initiative, our_problem}
    - daily  : [{"date": "2026-08-18", "count": n, "resolved": n}, ...]
               按日升序；trend_start/trend_end（"YYYY-MM-DD"）仅过滤该序列
    - regions: [{"region": "广东", "count": n}, ...] 按数量降序
    - types  : [{"issue_type": "硬件问题", "count": n, "resolved": n,
                 "unresolved": n}, ...] 按数量降序

    性能（P2，2026-08-28）：summary/daily/regions/types 全部 SQL 聚合
    （GROUP BY + SUM(CASE)），周期筛选 SQL 化（_append_cycle_where）——
    旧实现全表取回 7 字段 + Python 逐条聚合，100k 下每次弹窗打开数百毫秒。
    """
    conn = _conn()
    where, params = _build_where(keyword, issue_type, "")
    where, params = _append_cycle_where(where, params, cycle_start)

    # summary：单条 SQL 聚合
    row = conn.execute(
        "SELECT COUNT(*), SUM(resolved = '是'), SUM(is_initiative = '是'), "
        "SUM(is_our_problem = '是') FROM aftersale_records" + where,
        params).fetchone()
    n_all = int(row[0] or 0)
    n_resolved = int(row[1] or 0)
    n_init = int(row[2] or 0)
    n_our = int(row[3] or 0)
    rate = int(round(n_resolved * 100 / n_all)) if n_all else 0
    summary = {"total": n_all, "resolved": n_resolved,
               "unresolved": n_all - n_resolved, "rate": rate,
               "initiative": n_init, "our_problem": n_our}

    date_col = _RECORD_DATE_EXPR
    # 每日趋势：按记录日期聚合（occurred_at 优先回退 created_at），
    # trend_start/trend_end 仅作用于 daily 序列
    daily_where, daily_params = where, list(params)
    if trend_start:
        daily_where += (" AND " if daily_where else " WHERE ") + \
                       f"{date_col} >= ?"
        daily_params.append(str(trend_start))
    if trend_end:
        daily_where += (" AND " if daily_where else " WHERE ") + \
                       f"{date_col} <= ?"
        daily_params.append(str(trend_end))
    rows = conn.execute(
        f"SELECT {date_col} AS d, COUNT(*), SUM(resolved = '是') "
        f"FROM aftersale_records{daily_where} GROUP BY {date_col} "
        f"ORDER BY d ASC", daily_params).fetchall()
    daily = [{"date": str(d or ""), "count": int(c), "resolved": int(r or 0)}
             for d, c, r in rows if d]

    # 按地区聚合（数量降序；count 并列时按地区名稳定排序）
    rows = conn.execute(
        "SELECT region, COUNT(*) FROM aftersale_records" + where +
        " GROUP BY region ORDER BY COUNT(*) DESC, region ASC", params).fetchall()
    regions = [{"region": str(r or "").strip() or "未填地区", "count": int(c)}
               for r, c in rows]

    # 按问题类型聚合（数量降序；count 并列时按类型名稳定排序）
    rows = conn.execute(
        "SELECT issue_type, COUNT(*), SUM(resolved = '是') "
        "FROM aftersale_records" + where +
        " GROUP BY issue_type ORDER BY COUNT(*) DESC, issue_type ASC",
        params).fetchall()
    types = [{"issue_type": str(t or "").strip() or "未分类",
              "count": int(c), "resolved": int(rd or 0),
              "unresolved": int(c) - int(rd or 0)}
             for t, c, rd in rows]
    return {"summary": summary, "daily": daily,
            "regions": regions, "types": types}


def get_cycle_options() -> list:
    """周期下拉选项：库中记录实际归属的周期（按 occurred_at 动态计算，

    缺失回退 created_at），按起始日降序。仅返回库中确实存在数据的周期，
    不额外插入库中不存在的当前周期（当前周期无数据则不出现）。

    性能（P2，2026-08-28）：周期是记录日期的函数，故只需取 DISTINCT 日期
    （最多几百个唯一值）再逐个映射周期——旧实现全表取回 (occurred_at,
    created_at) 两列并在 Python 逐行算周期，100k 下每次面板加载数百毫秒。
    """
    from datetime import datetime
    conn = _conn()
    cur = conn.execute(
        f"SELECT DISTINCT {_RECORD_DATE_EXPR} AS d FROM aftersale_records")
    cycles = set()
    for (d,) in cur.fetchall():
        if not d:
            continue
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            continue
        c = cycle_start_of(dt)
        if c:
            cycles.add(c)
    cycles.discard(None)
    return sorted(cycles, reverse=True)


# ---------------- 上次填写记忆（重新打开面板时恢复填写人/解决人） ----------------
_LAST_CREATOR_KEY = "aftersale_last_creator"
_LAST_RESOLVER_KEY = "aftersale_last_resolver"


def _load_settings_str(key: str) -> str:
    """读取 settings.json 中指定字符串键（非空字符串返回，否则空串）"""
    import json
    from core.app_paths import get_app_dir
    try:
        path = os.path.join(get_app_dir(), "settings.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                v = json.load(f).get(key)
            if isinstance(v, str) and v.strip():
                return v.strip()
    except Exception:
        pass
    return ""


def load_last_creator() -> str:
    """上次填写的填写人（settings.json，无则空串）"""
    return _load_settings_str(_LAST_CREATOR_KEY)


def load_last_resolver() -> str:
    """上次填写的解决人（settings.json，无则空串）"""
    return _load_settings_str(_LAST_RESOLVER_KEY)


def save_last_people(creator: str, resolver: str) -> None:
    """记住本次填写的填写人/解决人，供下次打开面板恢复（非空才写）"""
    import json
    from core.app_paths import get_app_dir
    creator = (creator or "").strip()
    resolver = (resolver or "").strip()
    if not creator and not resolver:
        return
    try:
        path = os.path.join(get_app_dir(), "settings.json")
        data = {}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        if creator:
            data[_LAST_CREATOR_KEY] = creator
        if resolver:
            data[_LAST_RESOLVER_KEY] = resolver
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# S4（2026-09-06）：动态候选进程内缓存（写后失效）——get_field_candidates
# 是 4 条 GROUP BY 全表聚合（100k 内存库单次 94ms），录入页 showEvent 与
# 新增/编辑弹窗每次打开都触发，数据未变时纯属重复劳动。写操作成功后
# 置空失效（insert/update/delete/mark_resolved_batch/delete_records/import）。
_field_cands_cache: dict | None = None


def _invalidate_field_cands_cache() -> None:
    """写操作成功后置空候选缓存（下次 get_field_candidates 重建）"""
    global _field_cands_cache
    _field_cands_cache = None


def get_field_candidates() -> dict:
    """动态候选：问题/解决人/地区/填写人（按使用频次降序，各取前 60）

    命中进程内缓存返回深副本（列表可变，防调用方污染缓存）。
    """
    global _field_cands_cache
    if _field_cands_cache is not None:
        return {k: list(v) for k, v in _field_cands_cache.items()}
    conn = _conn()
    out = {}
    for key, field in (("problems", "problem"),
                       ("resolvers", "resolver"),
                       ("regions", "region"),
                       ("creators", "creator")):
        cur = conn.execute(
            f"SELECT {field}, COUNT(*) FROM aftersale_records "
            f"WHERE {field} != '' GROUP BY {field} "
            f"ORDER BY COUNT(*) DESC LIMIT 60")
        out[key] = [r[0] for r in cur.fetchall()]
    # 问题候选合并预置常见项（新库无历史数据时下拉不为空）
    if not out["problems"]:
        out["problems"] = ["主机没有开机", "遥控器没反应", "程序没了",
                           "不能扫码", "识别不了", "记分牌显示不出来"]
    # 填写人/解决人候选合并团队默认人员（恒置顶、去重、保持前 60）
    for key in ("resolvers", "creators"):
        seen = set()
        merged = []
        for name in list(_CREW_PRESET) + list(out[key]):
            if name and name not in seen:
                seen.add(name)
                merged.append(name)
        out[key] = merged[:60]
    _field_cands_cache = {k: list(v) for k, v in out.items()}
    return out


# ==================== 导出 / 导入 ====================

def _safe_sheet_title(title: str) -> str:
    r"""Excel sheet 名约束：≤31 字符且不含 : \ / ? * [ ]，非法字符转 -"""
    for ch in ':/\\?*[]':
        title = title.replace(ch, "-")
    return title[:31] or "未分类"


def _write_export_sheet(ws, rows):
    """写导出 sheet 公共逻辑：表头（加粗浅青底）+ 数据行 + 列宽 + 冻结首行

    周期列按发生时间动态归属，与列表筛选口径一致。
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9F2F4")
    for c, (_key, header) in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r, rec in enumerate(rows, 2):
        for c, (key, _h) in enumerate(_EXPORT_HEADERS, 1):
            val = rec.get(key) or ""
            if key == "cycle_start":
                cs = _record_cycle(rec.get("occurred_at"),
                                   rec.get("created_at"))
                val = cycle_label(cs) if cs else ""
            ws.cell(row=r, column=c, value=str(val))
    widths = (12, 24, 10, 8, 28, 30, 10, 30, 10, 12, 18, 12, 10, 16)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    ws.freeze_panes = "A2"


def _write_stats_sheet(wb, rows):
    """统计图表 sheet：类型×状态数据表 + 类型数量柱状图 + 解决状态饼图

    布局：A1:D{last} 数据表（含合计行），F2 起柱状图；
    数据表下方 A 列放 已解决/未解决 两行供饼图引用，饼图锚定 D 列。
    """
    if not rows:
        return
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, Reference

    ws = wb.create_sheet("统计图表")
    # 统计各类型 总数/已解决/未解决
    counts = {}
    for rec in rows:
        t = str(rec.get("issue_type") or "未分类")
        item = counts.setdefault(t, [0, 0, 0])
        item[0] += 1
        if rec.get("resolved") == "是":
            item[1] += 1
        else:
            item[2] += 1
    ordered = sorted(counts.items(), key=lambda kv: kv[1][0], reverse=True)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9F2F4")
    for c, h in enumerate(("类型", "数量", "已解决", "未解决"), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    total = total_yes = total_no = 0
    for r, (t, (n, yes, no)) in enumerate(ordered, 2):
        ws.cell(row=r, column=1, value=t)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=yes)
        ws.cell(row=r, column=4, value=no)
        total += n
        total_yes += yes
        total_no += no
    last = len(ordered) + 1  # 合计行号
    ws.cell(row=last, column=1, value="合计").font = header_font
    ws.cell(row=last, column=2, value=total)
    ws.cell(row=last, column=3, value=total_yes)
    ws.cell(row=last, column=4, value=total_no)

    # 柱状图：各问题类型数量（不含合计行）
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "各问题类型数量"
    bar.y_axis.title = "数量"
    bar.add_data(Reference(ws, min_col=2, max_col=2,
                           min_row=1, max_row=last - 1),
                 titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=2,
                                 max_row=last - 1))
    bar.width = 16
    bar.height = 9
    ws.add_chart(bar, "F2")

    # 饼图：已解决/未解决占比（数据表下方单独两行，避免混入类型行）
    pr = last + 2  # 饼图数据起始行
    ws.cell(row=pr - 1, column=1, value="状态").font = header_font
    ws.cell(row=pr - 1, column=2, value="数量").font = header_font
    ws.cell(row=pr, column=1, value="已解决")
    ws.cell(row=pr, column=2, value=total_yes)
    ws.cell(row=pr + 1, column=1, value="未解决")
    ws.cell(row=pr + 1, column=2, value=total_no)
    pie = PieChart()
    pie.title = "解决状态占比"
    pie.add_data(Reference(ws, min_col=2, min_row=pr - 1, max_row=pr + 1),
                 titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=pr, max_row=pr + 1))
    pie.width = 12
    pie.height = 9
    ws.add_chart(pie, f"D{pr}")


def export_xlsx(path: str, keyword: str = "", cycle_start: str = "",
                issue_type: str = "", resolved: str = "",
                is_initiative: str = "", is_our_problem: str = "") -> int:
    """按筛选条件导出全部记录（不分页）为 xlsx，返回导出条数

    表头与 售后问题汇总8月.xlsx 对齐，附加 填写时间/填写人/周期 三列。
    导出结构：
    - Sheet「售后记录」：全部记录（原样）
    - 每个有数据的问题类型一个 Sheet（如「硬件问题」），便于按类型分发
    - Sheet「统计图表」：类型数量柱状图 + 解决状态饼图
    """
    from openpyxl import Workbook

    conn = _conn()
    where, params = _build_where(keyword, issue_type, resolved,
                                 is_initiative, is_our_problem)
    # P2（2026-08-28）：周期筛选 SQL 化（与列表/统计同口径），
    # 导出前不再全量取回后 Python 过滤
    where, params = _append_cycle_where(where, params, cycle_start)
    cur = conn.execute(
        f"SELECT {', '.join(RECORD_FIELDS)} FROM aftersale_records"
        f"{where} ORDER BY id DESC", params)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    wb = Workbook()
    # 主 sheet：全部记录
    ws = wb.active
    ws.title = "售后记录"
    _write_export_sheet(ws, rows)

    # 按问题类型分类：预置枚举顺序在前，库中额外类型排后
    by_type = {}
    for rec in rows:
        by_type.setdefault(str(rec.get("issue_type") or "未分类"),
                           []).append(rec)
    for t in ISSUE_TYPES:
        if t in by_type:
            ws_t = wb.create_sheet(_safe_sheet_title(t))
            _write_export_sheet(ws_t, by_type.pop(t))
    for t, group in by_type.items():
        ws_t = wb.create_sheet(_safe_sheet_title(t))
        _write_export_sheet(ws_t, group)

    # 统计图表 sheet
    _write_stats_sheet(wb, rows)

    out_dir = os.path.dirname(os.path.abspath(path))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(path)
    return len(rows)


def parse_excel_rows(xlsx_path: str) -> tuple:
    """解析售后汇总 Excel，返回 (Excel 表头列表, 记录字典列表)；不写库

    与 import_excel_rows 共用同一解析逻辑，供导入预览与正式导入使用。
    规则：
    - 表头按中文名定位（类型/球房/桌号/地区/问题/发生原因/是否解决/解决方案/解决人/响应时间）
    - 类型列在 Excel 中为分组首行标记，逐行向下填充
    - 跳过全空行；是否解决空值默认「否」
    - created_at 填解析时间，creator 标记「Excel导入」，周期按当前配置归属
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]  # 取第一个工作表（Sheet1 正式表）

    # 表头定位：首行按中文名匹配列号
    header_map = {}
    excel_headers = []
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        if h:
            header_map[h] = c
            excel_headers.append(h)
    need = ("类型", "球房", "桌号", "地区", "问题")
    missing = [h for h in need if h not in header_map]
    if missing:
        raise ValueError(f"表头缺少必需列: {'、'.join(missing)}")

    def _val(row_idx, header):
        c = header_map.get(header)
        if not c:
            return ""
        v = ws.cell(row=row_idx, column=c).value
        return str(v or "").strip()

    now = datetime.now()
    created_at = now.strftime("%Y-%m-%d %H:%M:%S")
    occurred = created_at[:10]  # 历史导入无发生时间，默认取导入当日
    cycle = cycle_start_of(now)
    last_type = ""
    data = []
    for r in range(2, ws.max_row + 1):
        issue_type = _val(r, "类型") or last_type
        if _val(r, "类型"):
            last_type = issue_type
        room = _val(r, "球房")
        table_no = _val(r, "桌号")
        problem = _val(r, "问题")
        if not room and not table_no and not problem:
            continue  # 全空行跳过
        resolved = _val(r, "是否解决") or "否"
        if resolved not in ("是", "否"):
            resolved = "否"
        data.append({
            "created_at": created_at, "occurred_at": occurred,
            "creator": "Excel导入", "issue_type": issue_type,
            "table_no": table_no, "room_name": room,
            "region": _val(r, "地区"), "problem": problem,
            "cause": _val(r, "发生原因"), "resolved": resolved,
            "solution": _val(r, "解决方案"), "resolver": _val(r, "解决人"),
            "response_time": _val(r, "响应时间"), "cycle_start": cycle,
        })
    return excel_headers, data


def import_excel_rows(xlsx_path: str) -> int:
    """一次性导入历史 Excel（售后问题汇总 格式），返回导入条数

    解析规则见 parse_excel_rows；解析结果批量写入数据库。
    """
    _headers, data = parse_excel_rows(xlsx_path)
    if not data:
        return 0
    order = ("created_at", "occurred_at", "creator", "issue_type", "table_no",
             "room_name", "region", "problem", "cause", "resolved",
             "solution", "resolver", "response_time", "cycle_start")
    conn = _conn()
    conn.executemany(
        "INSERT INTO aftersale_records "
        "(created_at, occurred_at, creator, issue_type, table_no, room_name, "
        "region, problem, cause, resolved, solution, resolver, response_time, "
        "snk_code, device_code, cycle_start) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [tuple(rec[k] for k in order) + ("", "") for rec in data])
    conn.commit()
    _invalidate_field_cands_cache()
    return len(data)
