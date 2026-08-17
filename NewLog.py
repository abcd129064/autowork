# -*- coding: utf-8 -*-
"""斯诺克视频日志全自动整理工具（Excel 驱动）

按当天 Excel（问题/未复现/精度 三个 sheet）中的署名筛选记录，逐条把
视频、同名 log、daily 日志、detect.bin、日期文件夹从 videos/{球房}
归档到输出目录下的「球房 and 球房」总文件夹。可独立 CLI 运行，
也被 NewLogWorker 后台线程收编调用（传 target_name）。
"""
import os
import sys
import json
import shutil
import openpyxl
from datetime import datetime, timedelta

TIME_FORMAT = "%Y-%m-%d %H:%M:%S.%f"


# GUI 收编默认署名（与历史行为一致；GUI 调用方通过 target_name 参数覆盖）
DEFAULT_TARGET_NAME = "沈喆"

# 路径默认值（settings.json 缺失时回退到当前用户桌面）
_DEFAULT_VIDEOS_DIR = os.path.expanduser(r"~\Desktop\videos")
_DEFAULT_EXCEL_DIR = os.path.expanduser(r"~\Desktop\excel")
_DEFAULT_OUT_DIR = os.path.expanduser(r"~\Desktop")


def _load_settings() -> dict:
    """读取 settings.json（与主程序同目录，敏感字段透明解密）；失败返回 {}"""
    try:
        from core.app_paths import get_app_dir
        from core.secrets import decrypt_settings
        path = os.path.join(get_app_dir(), "settings.json")
        with open(path, "r", encoding="utf-8") as f:
            return decrypt_settings(json.load(f))
    except Exception:
        return {}


def parse_time_from_line(line):
    """解析日志行 [时间戳] 前缀；无括号或格式不对返回 None"""
    try:
        left = line.find('[')
        right = line.find(']')
        if left == -1 or right == -1:
            return None
        time_str = line[left + 1: right].strip()
        return datetime.strptime(time_str, TIME_FORMAT)
    except:
        return None


def get_log_start_end_time(log_path):
    """取日志首/末行时间戳作为时间范围（GBK 编码，容错返回 None）"""
    try:
        with open(log_path, "r", encoding="gbk", errors="ignore") as f:
            lines = [line.strip() for line in f if line.strip()]
        if not lines:
            return None, None
        start_t = parse_time_from_line(lines[0])
        end_t = parse_time_from_line(lines[-1])
        return start_t, end_t
    except:
        return None, None


def filter_daily_log(daily_path, expand_start, expand_end, save_path):
    """按时间窗口截取 daily 日志行另存为 UTF-8（带 BOM 供 Excel 直开）"""
    try:
        result = []
        with open(daily_path, "r", encoding="gbk", errors="ignore") as f:
            for line in f:
                try:
                    line = line.strip()
                    if not line:
                        continue
                    curr_t = parse_time_from_line(line)
                    if curr_t and expand_start <= curr_t <= expand_end:
                        result.append(line)
                except:
                    continue
        with open(save_path, "w", encoding="utf-8-sig") as f:
            f.write("\n".join(result))
        return True
    except:
        return False


def read_excel_data(excel_path, target_name=DEFAULT_TARGET_NAME):
    """按署名从三个 sheet 提取 (球房, 视频名) 列表，缺必要列的 sheet 跳过"""
    result = []
    all_rooms = set()
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheets_to_read = ["问题", "未复现", "精度"]

        for sheet_name in sheets_to_read:
            if sheet_name not in wb.sheetnames:
                print(f"  ⚠️ 工作表 '{sheet_name}' 不存在，跳过")
                continue

            ws = wb[sheet_name]

            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx

            if "署名" not in headers or "球房" not in headers or "视频名" not in headers:
                print(f"  ⚠️ 工作表 '{sheet_name}' 缺少必要的列（署名/球房/视频名），跳过")
                continue

            sig_col = headers["署名"]
            room_col = headers["球房"]
            video_col = headers["视频名"]

            for row_idx in range(2, ws.max_row + 1):
                try:
                    sig_cell = ws.cell(row_idx, sig_col).value
                    if sig_cell and sig_cell.strip() == target_name:
                        room_val = ws.cell(row_idx, room_col).value
                        video_val = ws.cell(row_idx, video_col).value

                        if not room_val or not video_val:
                            continue

                        room = str(room_val).strip()
                        video = str(video_val).strip()
                        result.append((room, video))
                        all_rooms.add(room)
                        print(f"  📋 从 '{sheet_name}' 提取: 球房={room}, 视频名={video}")
                except Exception as e:
                    continue

        wb.close()
        print(f"\n✅ 共提取 {len(result)} 条记录（署名：{target_name}）")
        print(f"✅ 涉及球桌: {', '.join(sorted(all_rooms))}")
        return result, all_rooms

    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return [], set()


def get_video_date_from_name(video_name):
    """视频名前 8 位数字（如 20260729_xxx）转成 YYYY-MM-DD，不合法返回 None"""
    try:
        date_part = video_name.split('_')[0]
        if len(date_part) == 8 and date_part.isdigit():
            year = date_part[0:4]
            month = date_part[4:6]
            day = date_part[6:8]
            return f"{year}-{month}-{day}"
    except:
        pass
    return None


def find_daily_log(table_path, date_str):
    """在球桌目录找含指定日期的 daily 开头的日志文件"""
    try:
        for f in os.listdir(table_path):
            if f.lower().startswith("daily") and date_str in f:
                return os.path.join(table_path, f)
    except:
        pass
    return None


def find_log_file(table_path, video_name):
    """找与视频同名的 .log 文件（不存在返回 None）"""
    try:
        log_file = f"{video_name}.log"
        log_path = os.path.join(table_path, log_file)
        if os.path.exists(log_path):
            return log_path
    except:
        pass
    return None


def get_today_excel_path():
    """当天 Excel 路径：{excel_dir}/{YYYY MM DD}.xlsx"""
    today = datetime.now()
    # 文件名用空格分隔（"2026 07 29.xlsx"），与台账模板约定一致
    date_str = today.strftime("%Y %m %d")
    excel_filename = f"{date_str}.xlsx"
    # Excel 目录来自 settings.json 的 newlog_excel_dir，缺失时回退桌面默认值
    excel_dir = str(_load_settings().get("newlog_excel_dir") or _DEFAULT_EXCEL_DIR)
    return os.path.join(excel_dir, excel_filename)


def main(target_name=None, interactive=False):
    """整理主流程：读 Excel → 逐条归档 → 汇总打印，返回输出总目录"""
    print("==================================================")
    print("     斯诺克视频日志自动整理工具（Excel驱动版）")
    print("        根据Excel署名筛选，自动归档")
    print("==================================================\n")

    # GUI 收编（NewLogWorker 后台线程）调用时传入 target_name；
    # 独立 CLI 运行时使用默认署名
    target_name = (str(target_name or "").strip() or DEFAULT_TARGET_NAME)
    
    # 配置路径（settings.json 驱动：videos_dir / newlog_excel_dir / newlog_out_dir）
    settings = _load_settings()
    work_root = str(settings.get("videos_dir") or _DEFAULT_VIDEOS_DIR)
    out_save_dir = str(settings.get("newlog_out_dir") or _DEFAULT_OUT_DIR)

    # 自动获取当天日期的Excel文件路径
    excel_path = get_today_excel_path()

    # 检查Excel文件是否存在
    if not os.path.exists(excel_path):
        print(f"❌ Excel文件不存在: {excel_path}")
        print(f"   请确保今天的Excel文件已放置在: {os.path.dirname(excel_path)}")
        print(f"   文件名格式应为: {datetime.now().strftime('%Y %m %d')}.xlsx")
        if interactive:
            input("\n按下回车键退出...")
        return

    print(f"📄 使用Excel文件: {os.path.basename(excel_path)}")
    print("📖 正在读取Excel文件...")
    records, all_rooms = read_excel_data(excel_path, target_name=target_name)

    if not records:
        print(f"❌ 未找到署名'{target_name}'的数据，程序退出。")
        if interactive:
            input("按下回车键退出...")
        return

    # 创建总文件夹，名字为所有球桌号用 " and " 连接
    folder_name = " and ".join(sorted(all_rooms))
    main_out_path = os.path.join(out_save_dir, folder_name)

    if not os.path.exists(main_out_path):
        os.makedirs(main_out_path)
        print(f"\n📁 创建总文件夹: {folder_name}")

    print(f"\n🚀 开始处理 {len(records)} 条记录...\n")

    success_count = 0
    fail_count = 0

    for idx, (room, video_name) in enumerate(records, 1):
        print(f"\n[{idx}/{len(records)}] 处理: 球房={room}, 视频名={video_name}")

        try:
            # 1. 构建球桌路径
            table_path = os.path.join(work_root, room)
            if not os.path.exists(table_path):
                print(f"  ❌ 球桌文件夹不存在: {table_path}")
                fail_count += 1
                continue

            # 2. 在总文件夹下创建球桌子文件夹
            table_out_path = os.path.join(main_out_path, room)
            if not os.path.exists(table_out_path):
                os.makedirs(table_out_path)
                print(f"  📁 创建子文件夹: {room}")

            # 3. 查找视频文件（在根目录）
            video_file = f"{video_name}.mp4"
            src_video = os.path.join(table_path, video_file)

            if not os.path.exists(src_video):
                print(f"  ❌ 视频文件不存在: {video_file}")
                fail_count += 1
                continue

            # 4. 复制视频文件
            dst_video = os.path.join(table_out_path, video_file)
            shutil.copy2(src_video, dst_video)
            print(f"  ✅ 视频复制完成: {video_file}")

            # 5. 查找并复制log文件
            src_log = find_log_file(table_path, video_name)
            if src_log:
                dst_log = os.path.join(table_out_path, f"{video_name}.log")
                shutil.copy2(src_log, dst_log)
                print(f"  ✅ log复制完成: {video_name}.log")
            else:
                print(f"  ⚠️ log文件不存在，跳过")

            # 6. 从视频名提取日期，查找并复制daily日志
            date_str = get_video_date_from_name(video_name)
            if date_str:
                daily_log = find_daily_log(table_path, date_str)
                if daily_log and os.path.exists(daily_log):
                    dst_daily = os.path.join(table_out_path, os.path.basename(daily_log))
                    shutil.copy2(daily_log, dst_daily)
                    print(f"  ✅ daily日志复制完成: {os.path.basename(daily_log)}")
                else:
                    print(f"  ⚠️ daily日志不存在，跳过")
            else:
                print(f"  ⚠️ 无法从视频名提取日期，跳过daily日志")

            # 7. 复制detect.bin
            src_bin = os.path.join(table_path, "detect.bin")
            if os.path.exists(src_bin):
                dst_bin = os.path.join(table_out_path, "detect.bin")
                shutil.copy2(src_bin, dst_bin)
                print(f"  ✅ detect.bin复制完成")
            else:
                print(f"  ⚠️ detect.bin不存在，跳过")

            # 8. 查找并复制日期文件夹（如果存在）
            if date_str:
                possible_date_folders = [date_str, date_str.replace("-", "")]
                for date_folder_name in possible_date_folders:
                    src_date_folder = os.path.join(table_path, date_folder_name)
                    if os.path.exists(src_date_folder) and os.path.isdir(src_date_folder):
                        dst_date_folder = os.path.join(table_out_path, date_folder_name)
                        if os.path.exists(dst_date_folder):
                            shutil.rmtree(dst_date_folder)
                        shutil.copytree(src_date_folder, dst_date_folder)
                        print(f"  ✅ 日期文件夹复制完成: {date_folder_name}")
                        break
                else:
                    print(f"  ⚠️ 日期文件夹不存在，跳过")

            success_count += 1
            print(f"  ✅ 处理完成!")

        except Exception as e:
            print(f"  ❌ 处理异常: {str(e)}")
            fail_count += 1
            continue

    print("\n" + "=" * 50)
    print(f"📊 处理完成！")
    print(f"   ✅ 成功: {success_count} 条")
    print(f"   ❌ 失败: {fail_count} 条")
    print(f"   📁 输出目录: {main_out_path}")
    print("=" * 50)

    if interactive:
        input("\n按下回车键关闭工具...")

    return main_out_path


if __name__ == "__main__":
    main(interactive=True)